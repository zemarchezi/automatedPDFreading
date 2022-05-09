#%%
from PIL import Image
from io import StringIO
import matplotlib.pyplot as plt
from generateLatexFile import *
import string
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# %%
def saveFigsGeo(image, outputpath, figname, crop):
    if crop:
        imgCropped = image[crop[0]:crop[1],crop[2]:crop[3]]
    else:
        imgCropped = image
    image.save(f'{outputpath}/{figname}.png')
    return f'{outputpath}/{figname}.png'


def extractFiguresTextGeomag(docPath, filename):

    wb = load_workbook(docPath+filename)
    

    textpt = []
    texten = []

    sheeten = wb['Plan3en']
    sheetpt = wb['Plan4pt']
    for r in range(sheeten.max_row):
        cell_obj = sheeten.cell(row = r+1, column = 1)
        texten.append(cell_obj.value)

    for r in range(sheetpt.max_row):
        cell_obj = sheetpt.cell(row = r+1, column = 1)
        textpt.append(cell_obj.value)
    # Put your sheet in the loader
    sheet = wb['Plan1']
    image_loader = SheetImageLoader(sheet)

    cellFigure = []
    for nn in range(1,80):
        for ll in list(string.ascii_uppercase):
            if image_loader.image_in(f'{ll}{nn}'):
                cellFigure.append(f"{ll}{nn}")

    

    return cellFigure, textpt, texten


def constructLatexFileGeomag(docPath, filename, outputFigure):
    cellFigure, tpt, ten = extractFiguresTextGeomag(docPath, filename)

    textpt = '\section{Geomagnetismo} \n \subsection{Responsável: Livia Ribeiro Alves} \n \n'
    texten = '\section{Geomagnetism} \n \subsection{Responsible: Livia Ribeiro Alves} \n \n'


    
    includeFigure = """\\begin{figure}[H]\n    
                        \\centering\n   
                             \\includegraphics[width=14cm]{./%s}\n
                        \\end{figure}\n
                     """
    wb = load_workbook(docPath+filename)
    sheet = wb['Plan1']
    image_loader = SheetImageLoader(sheet)
    for nn, im in enumerate(cellFigure):
        image = image_loader.get(im)
        outfigpath = saveFigsGeo(image, f"{outputFigure}", f'figureGeomag_{nn}', crop=False)
        texten += includeFigure % ('/'.join(outfigpath.split('/')[2:]))
        textpt += includeFigure % ('/'.join(outfigpath.split('/')[2:]))
    
    textpt += "\\begin{itemize} \n" 
    texten += "\\begin{itemize} \n" 
    for i in tpt:
        if not i.startswith('-'):
            textpt += "\\item " + i + "\n"
        else:
            textpt += "\\begin{itemize} \n \\item " + i + "\n \\end{itemize} \n"

    for i in ten:
        if not i.startswith('-'):
            texten += "\\item " + i + "\n"
        else:
            texten += "\\begin{itemize} \n \\item " + i + "\n \\end{itemize} \n"



    textpt += "\\end{itemize} \n"
    texten += "\\end{itemize} \n"
    # text += textSum


    return textpt, texten



PATH = './data/'

filename = 'Planilha_geom_briefing.xlsx'

#%%
# ssa = extractFiguresTextRoti(PATH, filename)
textpt, texten = constructLatexFileGeomag(PATH, filename, outputFigure='./latexText/figures')

generateLaTexFile(texten, EnPt=False, outputPath='./latexText', date='25/04/2022')

#%%
